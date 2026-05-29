from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Cliente, Servicio
import pandas as pd
from datetime import datetime
import os

router = APIRouter(prefix="/reportes", tags=["Reportes"])


#----------------------------
#Reporte 1: Clientes activos
#----------------------------
@router.get("/clientes-activos")
def reporte_clientes_activos(db: Session = Depends(get_db)):
    """
    Descarga Excel con lista de clientes activos
    
    Columnas:
    -ID
    -Nombre
    -Teléfono
    -Email
    -Fecha de registro
    """
    # Traer clientes activos
    clientes = db.query(Cliente).filter(Cliente.activo == True).all()
    
    # Convertir a lista de diccionarios
    datos = [
        {
            "ID": c.id,
            "Nombre": c.nombre,
            "Teléfono": c.telefono or "N/A",
            "Fecha de registro": c.creado_en.strftime("%Y-%m-%d")
        }
        for c in clientes
    ]
    # Si no hay datos, retornar tabla vacía
    if not datos:
        datos = [{"ID": "", "Nombre": "No hay clientes activos", "Teléfono": "", "Email": "", "Fecha de registro": ""}]
        
    # Crear DataFrame
    df = pd.DataFrame(datos)
    
    # Guardar a Excel
    archivo = "clientes_activos.xlsx"
    df.to_excel(archivo, sheet_name="Clientes Activos", index=False)
    
    #Devolver para descargar
    return FileResponse(
        archivo,
        media_type="application/vnd.ms-excel",
        filename=archivo
    )
    #-----------------------------
    # Reporte 2: Ingresos por mes
    #-----------------------------
    
@router.get("/ingresos-mes")
def reporte_ingresos_mes(db: Session = Depends(get_db)):
    """
    Descarga Excel con ingresos por mes
    
    Columnas:
    -Mes
    -Ingresos totales
    -Cantidad de servicios
    -Promedio por servicio
    """
        
    # Traer servicios
    servicios = db.query(Servicio).all()
        
    # Agrupar por mes
    ingresos_por_mes={}
    
    for servicio in servicios:
        mes = servicio.fecha.strftime("%Y-%m")
        
        if mes not in ingresos_por_mes:
            ingresos_por_mes[mes] = {
                "cantidad": 0,
                "total": 0.0
            }
        
        ingresos_por_mes[mes]["total"] += servicio.costo
        ingresos_por_mes[mes]["cantidad"] += 1
        
    # Convertir a tabla
    datos = [
        {
            "Mes": mes,
            "Ingresos totales": f"${round(info['total'], 2)}",
            "Cantidad de servicios": info["cantidad"],
            "Promedio por servicio": f"${round(info['total'] / info['cantidad'], 2)}"
        }
        for mes, info in sorted(ingresos_por_mes.items())
    ]
    if not datos:
        datos = [{"Mes": "Sin datos", "Ingresos totales": "$0.00", "Cantidad de servicios": 0, "Promedio por servicio": "$0.00"}]
        
    # DataFrame
    df = pd.DataFrame(datos)
    
    # Guardar
    archivo = "ingresos_mes.xlsx"
    df.to_excel(archivo, sheet_name="Ingresos", index=False)
    
    return FileResponse(
        archivo,
        media_type="application/vnd.ms-excel",
        filename=archivo
    )
    
#---------------------------------
# Reporte 3: Servicios facturados
#---------------------------------

@router.get("/servicios-facturados")
def reporte_servicios_facturados(db: Session = Depends(get_db)):
    """
    Descarga Excel con servicios facturados
    
    Columnas:
    -Cliente
    -Descripcion
    -Costo
    -Fecha
    -Total(Suma final)
    """
    
    # Traer servicios facturados
    servicios = db.query(Servicio).filter(Servicio.facturado == True).all()
    
    # Crear tabla
    datos = [
        {
            "Cliente": servicio.cliente.nombre,
            "Descripcion": servicio.descripcion,
            "Costo": f"${servicio.costo: .2f}",
            "Fecha": servicio.fecha.strftime("%Y-%m-%d"),
        }
        for servicio in servicios
    ]
    
    if not datos:
        datos = [{"Cliente": "Sin servicios facturados", "Descripcion": "", "Costo": "$0.00", "Fecha": ""}]
    # DataFrame
    df = pd.DataFrame(datos)
    
    # Agregar total
    if len(servicios) > 0:
        total = sum(s.costo for s in servicios)
        df.loc[len(df)] = ["Total", "", f"${total:.2f}", ""]
        
    # Guardar
    archivo = "servicios_facturados.xlsx"
    df.to_excel(archivo, sheet_name="Facturados", index=False)
    
    return FileResponse(
        archivo,
        media_type="application/vnd.ms-excel",
        filename=archivo
    )
    
    #-----------------------------------
    # Reporte 4: Servicios sin facturar
    #-----------------------------------
@router.get("/servicios-no-facturados")
def reporte_servicios_no_facturados(db: Session = Depends(get_db)):
    """
    Descarga Excel con servicios no facturados
    
    Columnas:
    -Cliente
    -Descripcion
    -Costo
    -Fecha
    -Total(Suma final)
    """
    
    # Traer servicios no facturados
    servicios = db.query(Servicio).filter(Servicio.facturado == False).all()
    
    # Crear tabla
    datos = [
        {
            "Cliente": servicio.cliente.nombre,
            "Descripcion": servicio.descripcion,
            "Costo": f"${servicio.costo: .2f}",
            "Fecha": servicio.fecha.strftime("%Y-%m-%d"),
        }
        for servicio in servicios
    ]
    
    if not datos:
        datos = [{"Cliente": "Sin servicios no facturados", "Descripcion": "", "Costo": "$0.00", "Fecha": ""}]
    # DataFrame
    df = pd.DataFrame(datos)
    
    # Agregar total
    if len(servicios) > 0:
        total = sum(s.costo for s in servicios)
        df.loc[len(df)] = ["Total", "", f"${total:.2f}", ""]
        
    # Guardar
    archivo = "servicios_no_facturados.xlsx"
    df.to_excel(archivo, sheet_name="No facturados", index=False)
    
    return FileResponse(
        archivo,
        media_type="application/vnd.ms-excel",
        filename=archivo
    )

# ─────────────────────────────────────────────────────────────────
# REPORTE 5: TOP CLIENTES (ANÁLISIS 80/20)
# ─────────────────────────────────────────────────────────────────
 
@router.get("/top-clientes")
def reporte_top_clientes(db: Session = Depends(get_db)):
    """
    Descarga Excel con top clientes por gasto total.
    Análisis 80/20: ¿cuáles son mis 20% de clientes que generan 80% de ingresos?
    
    Columnas:
    - Posición
    - Cliente
    - Total gastado
    - Cantidad de servicios
    - Porcentaje del total
    """
    
    # Traer servicios facturados
    servicios = db.query(Servicio).filter(Servicio.facturado == True).all()
    
    # Agrupar por cliente
    gasto_por_cliente = {}
    
    for servicio in servicios:
        nombre_cliente = servicio.cliente.nombre
        
        if nombre_cliente not in gasto_por_cliente:
            gasto_por_cliente[nombre_cliente] = {"total": 0.0, "cantidad": 0}
        
        gasto_por_cliente[nombre_cliente]["total"] += servicio.costo
        gasto_por_cliente[nombre_cliente]["cantidad"] += 1
    
    # Ordenar por gasto (de mayor a menor)
    clientes_ordenados = sorted(
        gasto_por_cliente.items(),
        key=lambda x: x[1]["total"],
        reverse=True
    )
    
    # Crear tabla con ranking
    total_general = sum(info["total"] for _, info in clientes_ordenados)
    
    datos = []
    for posicion, (cliente, info) in enumerate(clientes_ordenados, 1):
        porcentaje = (info["total"] / total_general * 100) if total_general > 0 else 0
        
        datos.append({
            "Posición": posicion,
            "Cliente": cliente,
            "Total gastado": f"${info['total']:.2f}",
            "Cantidad de servicios": info["cantidad"],
            "Porcentaje del total": f"{porcentaje:.1f}%"
        })
    
    if not datos:
        datos = [{"Posición": 1, "Cliente": "Sin datos", 
                  "Total gastado": "$0.00", "Cantidad de servicios": 0, 
                  "Porcentaje del total": "0.0%"}]
    
    # DataFrame
    df = pd.DataFrame(datos)
    
    # Guardar
    archivo = "top_clientes.xlsx"
    df.to_excel(archivo, sheet_name="Top Clientes", index=False)
    
    return FileResponse(
        archivo,
        media_type="application/vnd.ms-excel",
        filename=archivo
    )
 
 
# ─────────────────────────────────────────────────────────────────
# REPORTE 6: RESUMEN EJECUTIVO
# ─────────────────────────────────────────────────────────────────
 
@router.get("/resumen")
def reporte_resumen(db: Session = Depends(get_db)):
    """
    Descarga Excel con resumen ejecutivo del taller.
    
    Secciones:
    - KPIs (métricas clave)
    - Top 5 clientes
    - Últimos 10 servicios
    """
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    
    # Estilos
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # ═════════════════════════════════════════════
    # SECCIÓN 1: TÍTULO
    # ═════════════════════════════════════════════
    
    ws.merge_cells("A1:D1")
    titulo = ws["A1"]
    titulo.value = "RESUMEN EJECUTIVO DEL TALLER"
    titulo.font = titulo_font
    titulo.fill = titulo_fill
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # ═════════════════════════════════════════════
    # SECCIÓN 2: KPIs
    # ═════════════════════════════════════════════
    
    row = 3
    ws[f"A{row}"] = "MÉTRICAS CLAVE (KPIs)"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    
    row += 1
    
    # KPI 1: Total de clientes
    total_clientes = db.query(Cliente).count()
    ws[f"A{row}"] = "Total de clientes:"
    ws[f"B{row}"] = total_clientes
    row += 1
    
    # KPI 2: Clientes activos
    clientes_activos = db.query(Cliente).filter(Cliente.activo == True).count()
    ws[f"A{row}"] = "Clientes activos:"
    ws[f"B{row}"] = clientes_activos
    row += 1
    
    # KPI 3: Ingresos este mes
    hoy = datetime.utcnow()
    primero_mes = hoy.replace(day=1)
    
    servicios_mes = db.query(Servicio).filter(
        Servicio.fecha >= primero_mes,
        Servicio.facturado == True
    ).all()
    ingresos_mes = sum(s.costo for s in servicios_mes)
    
    ws[f"A{row}"] = "Ingresos este mes:"
    ws[f"B{row}"] = f"${ingresos_mes:.2f}"
    row += 1
    
    # KPI 4: Servicios pendientes
    pendientes = db.query(Servicio).filter(Servicio.facturado == False).count()
    ws[f"A{row}"] = "Servicios pendientes de facturar:"
    ws[f"B{row}"] = pendientes
    row += 1
    
    # KPI 5: Ingresos totales (histórico)
    todos_servicios = db.query(Servicio).filter(Servicio.facturado == True).all()
    ingresos_totales = sum(s.costo for s in todos_servicios)
    
    ws[f"A{row}"] = "Ingresos totales (histórico):"
    ws[f"B{row}"] = f"${ingresos_totales:.2f}"
    row += 2
    
    # ═════════════════════════════════════════════
    # SECCIÓN 3: TOP 5 CLIENTES
    # ═════════════════════════════════════════════
    
    ws[f"A{row}"] = "TOP 5 CLIENTES"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    # Headers
    ws[f"A{row}"] = "Cliente"
    ws[f"B{row}"] = "Total gastado"
    ws[f"A{row}"].font = header_font
    ws[f"B{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"B{row}"].fill = header_fill
    row += 1
    
    # Cálculo de top clientes
    gasto_por_cliente = {}
    for s in servicios:
        nombre = s.cliente.nombre
        if nombre not in gasto_por_cliente:
            gasto_por_cliente[nombre] = 0.0
        gasto_por_cliente[nombre] += s.costo
    
    clientes_top = sorted(
        gasto_por_cliente.items(),
        key=lambda x: x[1],
        reverse=True
    )[:5]
    
    for cliente, total in clientes_top:
        ws[f"A{row}"] = cliente
        ws[f"B{row}"] = f"${total:.2f}"
        row += 1
    
    row += 1
    
    # ═════════════════════════════════════════════
    # SECCIÓN 4: ÚLTIMOS 10 SERVICIOS
    # ═════════════════════════════════════════════
    
    ws[f"A{row}"] = "ÚLTIMOS 10 SERVICIOS"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    # Headers
    ws[f"A{row}"] = "Cliente"
    ws[f"B{row}"] = "Descripción"
    ws[f"C{row}"] = "Costo"
    ws[f"D{row}"] = "Facturado"
    
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = header_font
        ws[f"{col}{row}"].fill = header_fill
    
    row += 1
    
    # Datos
    ultimos_servicios = db.query(Servicio).order_by(Servicio.fecha.desc()).limit(10).all()
    for servicio in ultimos_servicios:
        ws[f"A{row}"] = servicio.cliente.nombre
        ws[f"B{row}"] = servicio.descripcion
        ws[f"C{row}"] = f"${servicio.costo:.2f}"
        ws[f"D{row}"] = "Sí" if servicio.facturado else "No"
        row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    
    # Guardar
    archivo = "resumen_ejecutivo.xlsx"
    wb.save(archivo)
    
    return FileResponse(
        archivo,
        media_type="application/vnd.ms-excel",
        filename=archivo
    )